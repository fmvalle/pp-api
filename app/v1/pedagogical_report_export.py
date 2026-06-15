"""Exportação PDF e Excel do relatório pedagógico individual."""

from __future__ import annotations

import re
import unicodedata
from datetime import datetime
from io import BytesIO
from typing import Any

_ACTION_LABEL = {
    "intervir": "Intervenção",
    "orientar": "Orientação",
    "desafiar": "Desafio",
}


def _strip_html(value: Any) -> str:
    if not value:
        return ""
    text = re.sub(r"<[^>]+>", " ", str(value))
    return re.sub(r"\s+", " ", text).strip()


def _fmt_pct(value: Any) -> str:
    if value is None:
        return "—"
    try:
        return f"{float(value):.1f}%"
    except (TypeError, ValueError):
        return str(value)


def _fmt_dt(value: Any) -> str:
    if not value:
        return "—"
    try:
        d = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return d.strftime("%d/%m/%Y")
    except Exception:
        return str(value)


def _safe_filename_part(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^\w\-]+", "-", ascii_only).strip("-").lower()
    return (slug[:40] or "export")


def pedagogical_export_filename(bundle: dict[str, Any], ext: str) -> str:
    student = str((bundle.get("student") or {}).get("name") or "aluno")
    assessment = str((bundle.get("assessment") or {}).get("title") or "avaliacao")
    return f"relatorio-{_safe_filename_part(student)}-{_safe_filename_part(assessment)[:30]}.{ext}"


def build_pedagogical_report_pdf_bytes(bundle: dict[str, Any]) -> bytes:
    """Monta PDF A4 do relatório pedagógico individual."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    assessment = bundle.get("assessment") or {}
    classroom = bundle.get("classroom") or {}
    student = bundle.get("student") or {}
    summary = bundle.get("summary") or {}
    reading = bundle.get("pedagogicalReading") or {}
    components = list(bundle.get("componentPerformance") or [])
    groups = list(bundle.get("questionGroups") or [])

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title="Relatório pedagógico individual",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="ReportTitle",
        parent=styles["Title"],
        fontSize=15,
        spaceAfter=10,
        textColor=colors.HexColor("#1a237e"),
    )
    section_style = ParagraphStyle(
        name="Section",
        parent=styles["Heading2"],
        fontSize=11,
        spaceBefore=10,
        spaceAfter=6,
        textColor=colors.HexColor("#283593"),
    )
    body_style = ParagraphStyle(
        name="Body",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
    )
    small_style = ParagraphStyle(
        name="Small",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
    )

    story: list[Any] = []
    story.append(Paragraph("Relatório pedagógico individual", title_style))
    meta_lines = [
        f"<b>Aluno:</b> {student.get('name') or '—'}",
        f"<b>Turma:</b> {classroom.get('name') or '—'}",
        f"<b>Escola:</b> {classroom.get('school') or '—'}",
        f"<b>Avaliação:</b> {assessment.get('title') or '—'}",
        f"<b>Data:</b> {_fmt_dt(assessment.get('date'))}",
    ]
    story.append(Paragraph("<br/>".join(meta_lines), body_style))
    story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Resumo", section_style))
    summary_rows = [
        ["Indicador", "Valor"],
        ["Total de questões", str(summary.get("totalQuestions") or "—")],
        ["Acertos", str(summary.get("correctAnswers") or "—")],
        ["Acerto do aluno", _fmt_pct(summary.get("accuracyPercentage"))],
        ["Média da turma", _fmt_pct(summary.get("classroomAverage"))],
        ["Média da escola", _fmt_pct(summary.get("schoolAverage"))],
        ["Média do sistema", _fmt_pct(summary.get("systemAverage"))],
    ]
    summary_table = Table(summary_rows, colWidths=[doc.width * 0.55, doc.width * 0.45])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8eaf6")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(summary_table)

    reading_text = str(reading.get("text") or "—")
    story.append(Paragraph("Leitura pedagógica", section_style))
    story.append(Paragraph(reading_text, body_style))

    if components:
        story.append(Paragraph("Componentes e ação pedagógica", section_style))
        comp_rows = [
            [
                "Componente",
                "Questões",
                "Acertos",
                "Acerto aluno",
                "Média ref.",
                "Variação (p.p.)",
                "Ação",
            ]
        ]
        for c in components:
            comp_rows.append(
                [
                    str(c.get("componentName") or "—"),
                    str(c.get("totalQuestions") or "—"),
                    str(c.get("correctAnswers") or "—"),
                    _fmt_pct(c.get("studentAccuracy")),
                    _fmt_pct(c.get("comparisonAverage")),
                    str(c.get("variationPercentagePoints") if c.get("variationPercentagePoints") is not None else "—"),
                    _ACTION_LABEL.get(str(c.get("pedagogicalAction") or ""), str(c.get("pedagogicalAction") or "—")),
                ]
            )
        comp_table = Table(
            comp_rows,
            colWidths=[
                doc.width * 0.22,
                doc.width * 0.08,
                doc.width * 0.08,
                doc.width * 0.12,
                doc.width * 0.12,
                doc.width * 0.12,
                doc.width * 0.12,
            ],
            repeatRows=1,
        )
        comp_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8eaf6")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(comp_table)

    if groups:
        story.append(Paragraph("Questão a questão", section_style))
        for group in groups:
            comp_name = str(group.get("componentName") or "Componente")
            area_name = str(group.get("areaName") or "")
            acc = _fmt_pct(group.get("accuracyPercentage"))
            subtitle = f"{comp_name}"
            if area_name and area_name != comp_name:
                subtitle += f" ({area_name})"
            subtitle += f" — acerto do aluno: {acc}"
            story.append(Paragraph(subtitle, body_style))
            q_rows = [
                [
                    "Nº",
                    "Habilidade",
                    "Gabarito",
                    "Resposta",
                    "Média escola",
                    "Média geral",
                    "Resp.",
                ]
            ]
            for q in group.get("questions") or []:
                skill = q.get("skillCode") or ""
                if q.get("skillDescription"):
                    skill = f"{skill} — {q.get('skillDescription')}" if skill else str(q.get("skillDescription"))
                q_rows.append(
                    [
                        str(q.get("questionNumber") or "—"),
                        skill or "—",
                        str(q.get("correctAnswer") or "—"),
                        str(q.get("studentAnswer") or "—"),
                        _fmt_pct(q.get("schoolAccuracyPercentage")),
                        _fmt_pct(q.get("systemAccuracyPercentage")),
                        str(q.get("totalResponses") or "—"),
                    ]
                )
            q_table = Table(
                q_rows,
                colWidths=[
                    doc.width * 0.06,
                    doc.width * 0.28,
                    doc.width * 0.08,
                    doc.width * 0.08,
                    doc.width * 0.12,
                    doc.width * 0.12,
                    doc.width * 0.08,
                ],
                repeatRows=1,
            )
            q_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 3),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                        ("TOPPADDING", (0, 0), (-1, -1), 2),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ]
                )
            )
            story.append(q_table)
            story.append(Spacer(1, 0.2 * cm))

    doc.build(story)
    return buf.getvalue()


def build_pedagogical_report_xlsx_bytes(bundle: dict[str, Any]) -> bytes:
    """Monta planilha Excel com resumo e questões (dados brutos)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    assessment = bundle.get("assessment") or {}
    classroom = bundle.get("classroom") or {}
    student = bundle.get("student") or {}
    summary = bundle.get("summary") or {}
    reading = bundle.get("pedagogicalReading") or {}
    components = list(bundle.get("componentPerformance") or [])
    groups = list(bundle.get("questionGroups") or [])

    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    bold = Font(bold=True)
    ws_resumo["A1"] = "Relatório pedagógico individual"
    ws_resumo["A1"].font = Font(bold=True, size=14)

    meta = [
        ("Aluno", student.get("name")),
        ("Turma", classroom.get("name")),
        ("Escola", classroom.get("school")),
        ("Avaliação", assessment.get("title")),
        ("Data", _fmt_dt(assessment.get("date"))),
        ("Total de questões", summary.get("totalQuestions")),
        ("Acertos", summary.get("correctAnswers")),
        ("Acerto do aluno (%)", summary.get("accuracyPercentage")),
        ("Média da turma (%)", summary.get("classroomAverage")),
        ("Média da escola (%)", summary.get("schoolAverage")),
        ("Média do sistema (%)", summary.get("systemAverage")),
        ("Leitura pedagógica", reading.get("text")),
    ]
    row = 3
    for label, value in meta:
        ws_resumo.cell(row=row, column=1, value=label).font = bold
        ws_resumo.cell(row=row, column=2, value=value if value is not None else "")
        row += 1

    row += 1
    comp_headers = [
        "Componente",
        "Área",
        "Questões",
        "Acertos",
        "Acerto aluno (%)",
        "Média referência (%)",
        "Variação (p.p.)",
        "Ação pedagógica",
    ]
    for col, header in enumerate(comp_headers, start=1):
        cell = ws_resumo.cell(row=row, column=col, value=header)
        cell.font = bold
    row += 1
    for c in components:
        ws_resumo.cell(row=row, column=1, value=c.get("componentName"))
        ws_resumo.cell(row=row, column=2, value=c.get("areaName"))
        ws_resumo.cell(row=row, column=3, value=c.get("totalQuestions"))
        ws_resumo.cell(row=row, column=4, value=c.get("correctAnswers"))
        ws_resumo.cell(row=row, column=5, value=c.get("studentAccuracy"))
        ws_resumo.cell(row=row, column=6, value=c.get("comparisonAverage"))
        ws_resumo.cell(row=row, column=7, value=c.get("variationPercentagePoints"))
        ws_resumo.cell(
            row=row,
            column=8,
            value=_ACTION_LABEL.get(str(c.get("pedagogicalAction") or ""), c.get("pedagogicalAction")),
        )
        row += 1

    ws_questoes = wb.create_sheet("Questões")
    q_headers = [
        "Aluno",
        "Turma",
        "Escola",
        "Avaliação",
        "Área",
        "Componente",
        "Nº questão",
        "Tipo",
        "Código habilidade",
        "Descrição habilidade",
        "Gabarito",
        "Resposta aluno",
        "Acertou",
        "Média escola (%)",
        "Média geral (%)",
        "Respondentes turma",
        "Enunciado",
    ]
    for col, header in enumerate(q_headers, start=1):
        cell = ws_questoes.cell(row=1, column=col, value=header)
        cell.font = bold

    q_row = 2
    student_name = student.get("name")
    classroom_name = classroom.get("name")
    school_name = classroom.get("school")
    assessment_title = assessment.get("title")
    for group in groups:
        area_name = group.get("areaName")
        comp_name = group.get("componentName")
        for q in group.get("questions") or []:
            ws_questoes.cell(row=q_row, column=1, value=student_name)
            ws_questoes.cell(row=q_row, column=2, value=classroom_name)
            ws_questoes.cell(row=q_row, column=3, value=school_name)
            ws_questoes.cell(row=q_row, column=4, value=assessment_title)
            ws_questoes.cell(row=q_row, column=5, value=area_name)
            ws_questoes.cell(row=q_row, column=6, value=comp_name)
            ws_questoes.cell(row=q_row, column=7, value=q.get("questionNumber"))
            ws_questoes.cell(row=q_row, column=8, value=q.get("questionType"))
            ws_questoes.cell(row=q_row, column=9, value=q.get("skillCode"))
            ws_questoes.cell(row=q_row, column=10, value=q.get("skillDescription"))
            ws_questoes.cell(row=q_row, column=11, value=q.get("correctAnswer"))
            ws_questoes.cell(row=q_row, column=12, value=q.get("studentAnswer"))
            is_correct = q.get("isCorrect")
            ws_questoes.cell(
                row=q_row,
                column=13,
                value=("Sim" if is_correct else "Não") if is_correct is not None else "",
            )
            ws_questoes.cell(row=q_row, column=14, value=q.get("schoolAccuracyPercentage"))
            ws_questoes.cell(row=q_row, column=15, value=q.get("systemAccuracyPercentage"))
            ws_questoes.cell(row=q_row, column=16, value=q.get("totalResponses"))
            ws_questoes.cell(row=q_row, column=17, value=_strip_html(q.get("description")))
            q_row += 1

    for sheet in (ws_resumo, ws_questoes):
        for col_idx in range(1, sheet.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 12
            for cell in sheet[letter]:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), 48))
            sheet.column_dimensions[letter].width = max_len

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
