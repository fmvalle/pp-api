"""Parse CSV/XLSX de respostas e proficiências TRI."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

QUESTION_COL_RE = re.compile(r"^Q(\d{3})$", re.IGNORECASE)
MAX_FILE_BYTES = 10 * 1024 * 1024

HEADER_ALIASES: dict[str, str] = {
    "codigo_cartao": "codigo_cartao",
    "codigo cartao": "codigo_cartao",
    "cartao": "codigo_cartao",
    "ra": "ra",
    "escola": "escola",
    "serie": "serie",
    "ano_detectado": "ano_detectado",
    "ano": "ano_detectado",
    "caderno": "caderno",
    "turma": "turma",
    "estudante": "estudante",
    "aluno": "estudante",
    "prof_lp": "prof_lp",
    "proflp": "prof_lp",
    "prof_mt": "prof_mt",
    "profmt": "prof_mt",
    "prof_ch": "prof_ch",
    "prof_cn": "prof_cn",
}


def _normalize_header(value: str) -> str:
    key = str(value or "").strip().lower()
    key = re.sub(r"\s+", " ", key)
    if key in HEADER_ALIASES:
        return HEADER_ALIASES[key]
    qm = QUESTION_COL_RE.match(key.upper().replace(" ", ""))
    if qm:
        return f"Q{qm.group(1)}"
    return key


def _normalize_row_keys(row: dict[str, Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for raw_key, raw_val in row.items():
        norm = _normalize_header(str(raw_key))
        if not norm:
            continue
        out[norm] = "" if raw_val is None else str(raw_val).strip()
    return out


def parse_tabular_bytes(content: bytes, filename: str) -> tuple[list[str], list[dict[str, str]], list[str]]:
    """Retorna (headers, rows normalizados, colunas extras detectadas)."""
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return _parse_xlsx(content)
    if lower.endswith(".csv"):
        return _parse_csv(content)
    raise ValueError("Extensão não suportada. Use .csv ou .xlsx")


def _detect_csv_delimiter(text: str) -> str:
    """Detecta ; vs , (comum em exportações BR)."""
    first_line = text.splitlines()[0] if text else ""
    if first_line.count(";") > first_line.count(","):
        return ";"
    if first_line.count(",") > 0:
        return ","
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=";,\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, str]], list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    delimiter = _detect_csv_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        raise ValueError("Arquivo CSV sem cabeçalho")
    headers = [_normalize_header(h) for h in reader.fieldnames if h]
    rows: list[dict[str, str]] = []
    for raw in reader:
        rows.append(_normalize_row_keys(raw))
    extras = _detect_extra_columns(headers, rows)
    return headers, rows, extras


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, str]], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Planilha XLSX vazia")
    iter_rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(iter_rows)
    except StopIteration as exc:
        raise ValueError("Planilha XLSX sem cabeçalho") from exc
    raw_headers = [str(c).strip() if c is not None else "" for c in header_row]
    headers = [_normalize_header(h) for h in raw_headers if h]
    rows: list[dict[str, str]] = []
    for values in iter_rows:
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        raw: dict[str, Any] = {}
        for idx, h in enumerate(raw_headers):
            if not h:
                continue
            val = values[idx] if idx < len(values) else None
            raw[h] = val
        rows.append(_normalize_row_keys(raw))
    extras = _detect_extra_columns(headers, rows)
    return headers, rows, extras


def _detect_extra_columns(headers: list[str], rows: list[dict[str, str]]) -> list[str]:
    found: set[str] = set()
    for h in headers:
        if h in ("prof_ch", "prof_cn"):
            found.add(h)
    for row in rows:
        for key in ("prof_ch", "prof_cn"):
            if row.get(key):
                found.add(key)
    return sorted(found)
